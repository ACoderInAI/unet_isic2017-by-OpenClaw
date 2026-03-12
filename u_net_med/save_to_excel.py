"""
Save training metrics from TensorBoard to Excel.
"""
import sys
from tensorboard.backend.event_processing import event_accumulator
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Path to the TensorBoard log
log_path = r""

# Load the event accumulator
ea = event_accumulator.EventAccumulator(log_path, size_guidance={'scalars': 0})
ea.Reload()

# Get epoch-level metrics
metrics = {}
for tag in ['epoch/train_loss', 'epoch/train_dice', 'epoch/val_loss', 'epoch/val_dice', 'lr']:
    if tag in ea.Tags()['scalars']:
        events = ea.Scalars(tag)
        metrics[tag.replace('epoch/', '')] = [e.value for e in events]

# Create DataFrame
df = pd.DataFrame(metrics)
df.insert(0, 'epoch', range(1, len(df) + 1))

# Reorder columns
cols = ['epoch', 'train_loss', 'train_dice', 'val_loss', 'val_dice', 'lr']
df = df[cols]

# Save to Excel
output_path = r"u_net_med\unet_isic2017.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Write main data sheet
    df.to_excel(writer, sheet_name='Epoch Metrics', index=False)
    
    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Epoch Metrics']
    
    # Format headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create summary sheet
    summary_data = {
        'Metric': ['Total Epochs', 'Best Validation Dice', 'Best Epoch', 'Final Validation Dice', 'Final Train Dice', 'Final LR'],
        'Value': [
            len(df),
            f"{df['val_dice'].max():.4f}",
            int(df.loc[df['val_dice'].idxmax(), 'epoch']),
            f"{df['val_dice'].iloc[-1]:.4f}",
            f"{df['train_dice'].iloc[-1]:.4f}",
            f"{df['lr'].iloc[-1]:.2e}"
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Format summary sheet
    summary_ws = writer.sheets['Summary']
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    for cell in summary_ws['A']:
        cell.alignment = Alignment(horizontal='left')
    for cell in summary_ws['B']:
        cell.alignment = Alignment(horizontal='center')
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 20

print(f"✅ Excel file saved: {output_path}")
print(f"📊 Total epochs: {len(df)}")
print(f"🏆 Best val dice: {df['val_dice'].max():.4f} at epoch {df.loc[df['val_dice'].idxmax(), 'epoch']}")
